"""Anexo de datos en Excel (.xlsx) con gráficos nativos, usando openpyxl.

Reutiliza la MISMA agregación que el PDF (`datos.py`) y la misma marca
(`branding.py`). Los gráficos son objetos nativos de Excel (BarChart/LineChart)
que referencian rangos de celdas, de modo que el usuario puede editarlos. Sin
disponibilidad (RN-1.2).
"""
import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import branding


def _hex(rgb) -> str:
    return "%02X%02X%02X" % rgb


_AZUL = _hex(branding.AZUL)
_AZUL_CLARO = _hex(branding.AZUL_CLARO)
_GRIS_SUAVE = _hex(branding.GRIS_SUAVE)

_F_TITULO = Font(name="Calibri", size=14, bold=True, color=_AZUL)
_F_SUB = Font(name="Calibri", size=10, color=_hex(branding.GRIS))
_F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_F_KPI_LBL = Font(name="Calibri", size=9, color=_hex(branding.GRIS))
_F_KPI_VAL = Font(name="Calibri", size=13, bold=True, color=_AZUL)
_FILL_HEAD = PatternFill("solid", fgColor=_AZUL)
_FILL_ALT = PatternFill("solid", fgColor=_GRIS_SUAVE)
_INT = "#,##0"


def _marca(ws, subtitulo: str):
    """Título + subtítulo de marca en las dos primeras filas de la hoja."""
    ws["A1"] = branding.TITULO
    ws["A1"].font = _F_TITULO
    ws["A2"] = subtitulo
    ws["A2"].font = _F_SUB


def _pie(ws, actor_email=None):
    """Pie de impresión: confidencialidad + sello de generación (RNF-1.15/1.16)."""
    ws.oddFooter.center.text = (branding.PIE_CONFIDENCIAL + "  |  "
                                + branding.sello_generacion(actor_email))
    ws.oddFooter.center.size = 8


def _encabezado(ws, fila: int, headers, anchos=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=j, value=h)
        c.font = _F_HEAD
        c.fill = _FILL_HEAD
        c.alignment = Alignment(horizontal="center")
    if anchos:
        for j, w in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w


def _filas(ws, fila0: int, filas, formatos=None):
    """Escribe filas con cebra; `formatos` = número de formato por columna."""
    for i, row in enumerate(filas):
        r = fila0 + i
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            if i % 2:
                c.fill = _FILL_ALT
            if formatos and j - 1 < len(formatos) and formatos[j - 1]:
                c.number_format = formatos[j - 1]
    return fila0 + len(filas)


def _kpis(ws, fila0: int, kpis):
    """Bloque de KPIs (etiqueta arriba, valor abajo) en columnas de a pares."""
    r = fila0
    for i, (label, value, fmt) in enumerate(kpis):
        col = 1 + (i % 3) * 2
        lc = ws.cell(row=r, column=col, value=label)
        lc.font = _F_KPI_LBL
        vc = ws.cell(row=r + 1, column=col, value=value)
        vc.font = _F_KPI_VAL
        if fmt:
            vc.number_format = fmt
        if i % 3 == 2:
            r += 3
    return r + 3


def _bar(ws, titulo, anchor, data_ref, cats_ref, alto=7.5, ancho=18):
    ch = BarChart()
    ch.type = "col"
    ch.title = titulo
    ch.legend = None
    ch.height = alto
    ch.width = ancho
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    if ch.series:
        ch.series[0].graphicalProperties.solidFill = _AZUL
    ws.add_chart(ch, anchor)


def _line(ws, titulo, anchor, data_ref, cats_ref, alto=7.5, ancho=22):
    ch = LineChart()
    ch.title = titulo
    ch.legend = None
    ch.height = alto
    ch.width = ancho
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    if ch.series:
        ch.series[0].graphicalProperties.line.solidFill = _AZUL
        ch.series[0].graphicalProperties.line.width = 20000
    ws.add_chart(ch, anchor)


def _guardar(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hoja_sin_datos(wb, subtitulo, year, mes):
    ws = wb.active
    ws.title = "Resumen"
    _marca(ws, subtitulo)
    ws["A4"] = f"Sin cargas para {branding.nombre_mes(mes)} {year}"
    ws["A4"].font = _F_KPI_VAL
    ws["A5"] = "No hay escaneos registrados para este período."
    ws["A5"].font = _F_SUB
    ws.column_dimensions["A"].width = 40


# ══════════════════════════════════════════════════════════════
def informe_puerto(datos: dict, nombre: str, year: int, mes: int,
                   actor_email=None) -> bytes:
    sub = f"{nombre} · {branding.nombre_mes(mes)} {year}"
    wb = Workbook()

    if datos.get("sin_datos"):
        _hoja_sin_datos(wb, sub, year, mes)
        return _guardar(wb)

    # ── Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    _marca(ws, sub)
    _pie(ws, actor_email)
    v = datos["variacion"]
    tz = datos["trazabilidad"]
    kpis = [
        ("Total de escaneos", datos["total"], _INT),
        ("Promedio diario", datos["promedio_diario"], _INT),
        ("Días activos", datos["dias_activos"], _INT),
        ("Operadores", datos["operadores_distintos"], _INT),
        ("Contenedores", tz["contenedores"], _INT),
        ("Contenedores válidos", tz["contenedores_validos"], _INT),
        ("Placas", tz["placas"], _INT),
        ("Variación mensual %", v["mensual"]["delta_pct"], "0.0"),
        ("Variación interanual %", v["interanual"]["delta_pct"], "0.0"),
    ]
    _kpis(ws, 4, kpis)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 20

    # ── Serie diaria ──
    ws = wb.create_sheet("Serie diaria")
    _pie(ws, actor_email)
    _encabezado(ws, 1, ["Día", "Escaneos"], anchos=[10, 14])
    fin = _filas(ws, 2, [[d["dia"], d["total"]] for d in datos["serie_diaria"]],
                 formatos=[None, _INT])
    data = Reference(ws, min_col=2, min_row=1, max_row=fin - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=fin - 1)
    _bar(ws, "Escaneos por día", "D2", data, cats)

    # ── Distribución por hora ──
    ws = wb.create_sheet("Por hora")
    _pie(ws, actor_email)
    _encabezado(ws, 1, ["Hora", "Escaneos"], anchos=[10, 14])
    fin = _filas(ws, 2, [[f"{h['hora']:02d}", h["total"]] for h in datos["horaria"]],
                 formatos=[None, _INT])
    data = Reference(ws, min_col=2, min_row=1, max_row=fin - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=fin - 1)
    _bar(ws, "Distribución por hora del día", "D2", data, cats)

    # ── Operadores ──
    if datos["operadores"]:
        ws = wb.create_sheet("Operadores")
        _pie(ws, actor_email)
        _encabezado(ws, 1, ["Operador", "Escaneos"], anchos=[32, 14])
        fin = _filas(ws, 2, [[o["nombre"], o["total"]] for o in datos["operadores"]],
                     formatos=[None, _INT])
        data = Reference(ws, min_col=2, min_row=1, max_row=fin - 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=fin - 1)
        _bar(ws, "Productividad por operador", "D2", data, cats)

    # ── Alertas ──
    if datos["alertas"]:
        ws = wb.create_sheet("Alertas")
        _pie(ws, actor_email)
        _encabezado(ws, 1, ["Tipo", "Severidad", "Fecha", "Mensaje"],
                    anchos=[18, 14, 14, 60])
        _filas(ws, 2, [[a["tipo"], a["severidad"], branding.fecha_bogota(a["fecha"]),
                        a["mensaje"]] for a in datos["alertas"]])

    return _guardar(wb)


def informe_nacional(datos: dict, year: int, mes: int, actor_email=None) -> bytes:
    sub = f"Consolidado nacional · {branding.nombre_mes(mes)} {year}"
    wb = Workbook()

    if datos.get("sin_datos"):
        _hoja_sin_datos(wb, sub, year, mes)
        return _guardar(wb)

    # ── Resumen nacional ──
    ws = wb.active
    ws.title = "Resumen"
    _marca(ws, sub)
    _pie(ws, actor_email)
    v = datos["variacion"]
    kpis = [
        ("Escaneos nacionales", datos["total_nacional"], _INT),
        ("Promedio diario", datos["promedio_diario"], _INT),
        ("Puertos con datos", datos["puertos_con_dato"], _INT),
        ("Operadores", datos["operadores_distintos"], _INT),
        ("Contenedores", datos["contenedores"], _INT),
        ("Variación mensual %", v["mensual"]["delta_pct"], "0.0"),
        ("Variación interanual %", v["interanual"]["delta_pct"], "0.0"),
    ]
    _kpis(ws, 4, kpis)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 20

    # ── Ranking ──
    ws = wb.create_sheet("Ranking")
    _pie(ws, actor_email)
    _encabezado(ws, 1, ["Posición", "Puerto", "Escaneos", "Cuota %"],
                anchos=[10, 26, 14, 12])
    fin = _filas(ws, 2, [[r["posicion"], r["nombre"], r["total"], r["cuota_pct"]]
                         for r in datos["ranking"]],
                 formatos=[None, None, _INT, "0.0"])
    data = Reference(ws, min_col=3, min_row=1, max_row=fin - 1)
    cats = Reference(ws, min_col=2, min_row=2, max_row=fin - 1)
    _bar(ws, "Ranking de puertos por escaneos", "F2", data, cats)

    # ── Detalle por puerto ──
    ws = wb.create_sheet("Por puerto")
    _pie(ws, actor_email)
    _encabezado(ws, 1, ["Puerto", "Escaneos", "Prom/día", "Operadores",
                        "Contenedores", "Placas", "Alertas"],
                anchos=[26, 14, 12, 12, 14, 12, 10])
    _filas(ws, 2, [[p["nombre"], p["total"], p["promedio_diario"], p["operadores"],
                    p["contenedores"], p["placas"], p["alertas"]]
                   for p in datos["por_puerto"]],
           formatos=[None, _INT, _INT, _INT, _INT, _INT, _INT])

    # ── Serie anual ──
    ws = wb.create_sheet("Serie anual")
    _pie(ws, actor_email)
    _encabezado(ws, 1, ["Mes", "Escaneos"], anchos=[12, 14])
    fin = _filas(ws, 2, [[branding.MESES[x["mes"] - 1], x["total"]]
                         for x in datos["serie_anual"]],
                 formatos=[None, _INT])
    data = Reference(ws, min_col=2, min_row=1, max_row=fin - 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=fin - 1)
    _line(ws, f"Evolución mensual del año {year}", "D2", data, cats)

    return _guardar(wb)
