"""Anexo Excel (`reportes/excel.py`) vía los endpoints: estructura (hojas +
gráficos, re-abriendo con openpyxl), permisos por alcance, período vacío, auditoría."""
import io

from openpyxl import load_workbook

import database
from models import EscaneosDiarios

XLSX_MIME = "spreadsheetml"   # substring del content-type de .xlsx


def _seed(puerto_id, year, mes, n=5):
    db = database.SessionLocal()
    try:
        for dia in range(1, n + 1):
            db.add(EscaneosDiarios(puerto_id=puerto_id, year=year, mes=mes,
                                   dia=dia, total=100 + dia))
        db.commit()
    finally:
        db.close()


def _wb(content):
    return load_workbook(io.BytesIO(content))


def _assert_xlsx(r):
    assert r.status_code == 200, r.text
    assert XLSX_MIME in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    assert ".xlsx" in r.headers["content-disposition"]


# ── Puerto ──────────────────────────────────────────────────
def test_excel_puerto_ok(client, admin):
    _seed(0, 2026, 6)
    r = client.get("/reportes/excel/puerto/0/2026/6")
    _assert_xlsx(r)
    wb = _wb(r.content)
    assert {"Resumen", "Serie diaria", "Por hora"} <= set(wb.sheetnames)
    assert len(wb["Serie diaria"]._charts) >= 1      # gráfico nativo embebido


def test_excel_puerto_sin_sesion_401(client):
    assert client.get("/reportes/excel/puerto/0/2026/6").status_code == 401


def test_excel_puerto_periodo_vacio(client, admin):
    r = client.get("/reportes/excel/puerto/0/2026/1")
    _assert_xlsx(r)
    wb = _wb(r.content)
    assert wb.sheetnames == ["Resumen"]              # solo aviso "sin cargas"


def test_excel_puerto_feeder_su_puerto_ok(client, admin, feeder):
    _seed(0, 2026, 6)
    _assert_xlsx(feeder.get("/reportes/excel/puerto/0/2026/6"))


def test_excel_puerto_feeder_otro_puerto_403(client, admin, feeder):
    _seed(2, 2026, 6)
    assert feeder.get("/reportes/excel/puerto/2/2026/6").status_code == 403


def test_excel_puerto_ultimo(client, admin):
    _seed(0, 2026, 6)
    _seed(0, 2026, 7, n=3)
    r = client.get("/reportes/excel/puerto/0")
    _assert_xlsx(r)
    assert "2026-07" in r.headers["content-disposition"]


# ── Nacional ────────────────────────────────────────────────
def test_excel_nacional_admin_ok(client, admin):
    _seed(0, 2026, 6)
    _seed(1, 2026, 6)
    r = client.get("/reportes/excel/nacional/2026/6")
    _assert_xlsx(r)
    wb = _wb(r.content)
    assert {"Resumen", "Ranking", "Por puerto", "Serie anual"} <= set(wb.sheetnames)
    assert len(wb["Ranking"]._charts) >= 1
    assert len(wb["Serie anual"]._charts) >= 1


def test_excel_nacional_feeder_403(client, admin, feeder):
    _seed(0, 2026, 6)
    assert feeder.get("/reportes/excel/nacional/2026/6").status_code == 403
    assert feeder.get("/reportes/excel/nacional").status_code == 403


# ── Auditoría ───────────────────────────────────────────────
def test_excel_genera_auditoria(client, admin):
    _seed(0, 2026, 6)
    client.get("/reportes/excel/puerto/0/2026/6")
    eventos = client.get("/api/audit?limit=20").json()
    reporte = [e for e in eventos if e["accion"] == "reporte_excel"]
    assert reporte and reporte[0]["entidad"] == "reporte"
    assert client.get("/api/audit/verify").json()["ok"] is True
