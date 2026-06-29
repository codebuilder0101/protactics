"""
PROTACTICS — Backend API
FastAPI + SQLAlchemy + PostgreSQL (SQLite para desarrollo local)
"""
import io
import os
import calendar
from datetime import datetime, date
from typing import Optional

import openpyxl
import xlrd
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from database import get_db, init_db, SessionLocal
from models import (Puerto, EscaneosDiarios, EscaneosHorarios,
                    Operadores, Disponibilidad, ArchivosCargados, User, AuditLog,
                    Alerta, SLA, Infraccion, VentanaMantenimiento,
                    EscaneoFila, IndiceIdentificador)
from parsers import parse_file, detect_format
import identificadores
from routing import route_file, detect_period, detect_port
from audit import record_audit, verify_chain
from auth import (router as auth_router, get_current_user, require_admin,
                  user_from_token, COOKIE_NAME, can_view_port, can_upload_port,
                  can_manage_alerts, allowed_port_ids, ROLE_ADMIN, seed_demo_users)
import mantenimiento as mant
import sla as sla_engine
import anomalies

# ── App ────────────────────────────────────────────────────
app = FastAPI(title="PROTACTICS API", version="1.0.0")

# CORS restringido + credenciales (cookies). Un "*" es incompatible con
# cookies en el navegador, por eso se fija el/los origen(es) del frontend.
FRONTEND_ORIGINS = [o.strip() for o in os.getenv(
    "FRONTEND_ORIGIN", "http://localhost:8000,http://127.0.0.1:8000"
).split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=FRONTEND_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth_router)

# Registro público de cuentas. Deshabilitado por defecto (sitio privado).
# Para reactivarlo: define la variable de entorno REGISTRATION_ENABLED=true.
REGISTRATION_ENABLED = os.getenv("REGISTRATION_ENABLED", "false").lower() in ("1", "true", "yes")

# Carga masiva: tope de archivos por lote y tamaño máximo por archivo.
MAX_BULK_FILES = int(os.getenv("MAX_BULK_FILES", "100"))
MAX_FILE_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(15 * 1024 * 1024)))  # 15 MB


@app.on_event("startup")
def startup():
    # Asegura que los logs INFO de protactics.* aparezcan en los logs de Railway.
    import logging
    logging.basicConfig(level=logging.INFO)
    init_db()
    # Sembrar usuarios de demostración (uno por perfil) para pruebas.
    if os.getenv("SEED_DEMO_USERS", "true").lower() in ("1", "true", "yes"):
        db = SessionLocal()
        try:
            seed_demo_users(db)
        finally:
            db.close()


# ── Frontend (páginas) ─────────────────────────────────────
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")

# Recursos estáticos del frontend (favicon, imágenes, etc.) bajo /assets
app.mount(
    "/assets",
    StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")),
    name="assets",
)


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse(os.path.join(FRONTEND_DIR, "assets", "favicon.png"))


def _page(name: str) -> FileResponse:
    return FileResponse(
        os.path.join(FRONTEND_DIR, name),
        headers={"Cache-Control": "no-store"},  # evita ver páginas cacheadas tras logout
    )


@app.get("/")
def landing():
    """Página de aterrizaje pública (misión y propósito)."""
    return _page("landing.html")


@app.get("/login")
def login_page(request: Request, db: Session = Depends(get_db)):
    if user_from_token(db, request.cookies.get(COOKIE_NAME)):
        return RedirectResponse("/dashboard", status_code=302)
    return _page("login.html")


@app.get("/register")
def register_page(request: Request, db: Session = Depends(get_db)):
    if not REGISTRATION_ENABLED:
        return RedirectResponse("/login", status_code=302)
    if user_from_token(db, request.cookies.get(COOKIE_NAME)):
        return RedirectResponse("/dashboard", status_code=302)
    return _page("register.html")


@app.get("/dashboard")
def dashboard_page(request: Request, db: Session = Depends(get_db)):
    if not user_from_token(db, request.cookies.get(COOKIE_NAME)):
        return RedirectResponse("/login", status_code=302)
    return _page("index.html")


@app.get("/admin")
def admin_page(request: Request, db: Session = Depends(get_db)):
    user = user_from_token(db, request.cookies.get(COOKIE_NAME))
    if not user:
        return RedirectResponse("/login", status_code=302)
    if user.role != ROLE_ADMIN:
        return RedirectResponse("/dashboard", status_code=302)
    return _page("admin.html")


@app.get("/admin/approvals")
def approvals_page(request: Request, db: Session = Depends(get_db)):
    user = user_from_token(db, request.cookies.get(COOKIE_NAME))
    if not user:
        return RedirectResponse("/login", status_code=302)
    if user.role != ROLE_ADMIN:
        return RedirectResponse("/dashboard", status_code=302)
    return _page("approvals.html")


# ── Configuración pública del frontend ─────────────────────
@app.get("/api/config")
def get_config():
    """Flags que el frontend consulta (p. ej. si el registro está habilitado)."""
    return {"registration_enabled": REGISTRATION_ENABLED}


# ── Lista pública de puertos (solo nombres) para el registro ─
@app.get("/api/puertos/public")
def get_puertos_public(db: Session = Depends(get_db)):
    puertos = db.query(Puerto).order_by(Puerto.id).all()
    return [{"id": p.id, "nombre_corto": p.nombre_corto, "nombre": p.nombre}
            for p in puertos]


# ── MESES ──────────────────────────────────────────────────
MONTHS = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
          "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ── HELPERS ────────────────────────────────────────────────
# Marcadores de columnas que solo aparecen en la hoja de DETALLE (escaneos fila
# a fila). Sirven para elegir esa hoja en libros con varias hojas (p. ej. una de
# detalle "Imágenes de Escaneo" y otra de resumen "Estadísticas de Escaneo").
DETAIL_SHEET_MARKERS = (
    "Fecha de creaci", "Scan Date", "Escaneos Individuales", "Miniatura",
    "ID del Contenedor", "User Name", "Nombre de Usuario",
    "Estado de flujo de trabajo",
)


def _sheet_score(rows: list) -> int:
    """Cuenta marcadores de hoja de detalle en las primeras filas."""
    score = 0
    for row in rows[:26]:
        for c in row:
            if c is not None:
                s = str(c)
                score += sum(1 for m in DETAIL_SHEET_MARKERS if m in s)
    return score


def read_excel_rows(content: bytes, filename: str) -> list:
    """Lee XLS o XLSX y retorna lista de listas (header:1).

    Si el libro tiene varias hojas, elige la hoja de DETALLE (escaneos fila a
    fila) en lugar de una hoja de resumen, comparando marcadores de columnas
    conocidas. Para libros de una sola hoja el comportamiento no cambia.
    """
    if filename.lower().endswith(".xls"):
        book = xlrd.open_workbook(file_contents=content)
        best, best_score = book.sheet_by_index(0), -1
        for sh in book.sheets():
            head = [sh.row_values(i) for i in range(min(sh.nrows, 26))]
            score = _sheet_score(head)
            if score > best_score:
                best, best_score = sh, score
        return [best.row_values(i) for i in range(best.nrows)]
    else:
        # 1ª pasada: puntuar las cabeceras de cada hoja (barata, ~26 filas).
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        scores = {}
        for ws in wb.worksheets:
            head = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                head.append(list(row))
                if i >= 25:
                    break
            scores[ws.title] = _sheet_score(head)
        wb.close()
        best_title = max(scores, key=scores.get) if scores else None

        # 2ª pasada: leer completa la hoja elegida.
        wb2 = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb2[best_title] if best_title is not None else wb2.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb2.close()
        return rows


def rows_to_dicts(rows: list) -> list[dict]:
    """Convierte lista de arrays a lista de dicts usando la primera fila como header."""
    if not rows:
        return []
    if isinstance(rows[0], dict):
        return rows
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return result


def compute_availability(daily: dict, year: int, mes: int) -> float:
    """Estima la disponibilidad del servicio a partir de los días operativos.

    Disponibilidad = (días con actividad de escaneo / días del mes) × 100.
    Para el mes en curso se usan los días transcurridos hasta hoy.
    Devuelve un porcentaje realista entre 0 y 100 (1 decimal).
    """
    days_active = len([d for d, t in daily.items() if t])
    if days_active == 0:
        return 0.0
    days_in_month = calendar.monthrange(year, mes)[1]
    today = datetime.utcnow()
    eff_days = today.day if (year == today.year and mes == today.month) else days_in_month
    eff_days = max(eff_days, days_active)  # nunca dividir por menos que los días activos
    return round(min(100.0, days_active / eff_days * 100.0), 1)


def _by_day(data: dict) -> dict:
    """Obtiene el desglose por día del parser. Si por algún motivo no viene,
    lo reconstruye desde los agregados (queda todo bajo el día reportado)."""
    if data.get("by_day"):
        return data["by_day"]
    # Reconstrucción de respaldo: un único día (el primero de 'daily').
    by_day = {}
    for dia, total in data.get("daily", {}).items():
        by_day[int(dia)] = {"total": int(total),
                            "hourly": data.get("hourly", {}),
                            "operators": data.get("operators", {})}
    return by_day


def save_parsed_data(db: Session, puerto_id: int, year: int, mes: int,
                     data: dict, filename: str):
    """Acumula los datos del reporte. Solo reemplaza los DÍAS presentes en este
    archivo, de modo que los reportes diarios se suman en el mes sin borrar los
    días ya cargados. Volver a subir un día corrige solo ese día."""

    by_day = _by_day(data)

    for dia, info in by_day.items():
        dia = int(dia)
        # Reemplazar SOLO este día (acumulación + corrección idempotente).
        db.query(EscaneosDiarios).filter_by(
            puerto_id=puerto_id, year=year, mes=mes, dia=dia).delete()
        db.query(EscaneosHorarios).filter_by(
            puerto_id=puerto_id, year=year, mes=mes, dia=dia).delete()
        db.query(Operadores).filter_by(
            puerto_id=puerto_id, year=year, mes=mes, dia=dia).delete()

        db.add(EscaneosDiarios(puerto_id=puerto_id, year=year, mes=mes,
                               dia=dia, total=int(info["total"])))
        for hora, total in info.get("hourly", {}).items():
            db.add(EscaneosHorarios(puerto_id=puerto_id, year=year, mes=mes,
                                    dia=dia, hora=int(hora), total=int(total)))
        for nombre, total in info.get("operators", {}).items():
            db.add(Operadores(puerto_id=puerto_id, year=year, mes=mes,
                              dia=dia, nombre=str(nombre), total=int(total)))

    db.flush()

    # Total acumulado del mes (todos los días ya guardados).
    month_total = db.query(func.coalesce(func.sum(EscaneosDiarios.total), 0))\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).scalar() or 0
    month_daily = {d.dia: d.total for d in db.query(EscaneosDiarios)
                   .filter_by(puerto_id=puerto_id, year=year, mes=mes).all()}

    # Registro de archivo: refleja el total ACUMULADO del mes.
    existing = db.query(ArchivosCargados).filter_by(
        puerto_id=puerto_id, year=year, mes=mes).first()
    if existing:
        existing.nombre_archivo = filename
        existing.formato = data["format"]
        existing.total_escaneos = month_total
        existing.cargado_en = datetime.utcnow()
    else:
        db.add(ArchivosCargados(
            puerto_id=puerto_id, year=year, mes=mes,
            nombre_archivo=filename, formato=data["format"],
            total_escaneos=month_total
        ))

    # Disponibilidad estimada sobre el mes ACUMULADO. Solo si no hay valor manual.
    disp = db.query(Disponibilidad).filter_by(
        puerto_id=puerto_id, year=year, mes=mes).first()
    auto = compute_availability(month_daily, year, mes)
    if disp is None:
        db.add(Disponibilidad(puerto_id=puerto_id, year=year, mes=mes, valor=auto))
    elif disp.valor is None:
        disp.valor = auto

    db.commit()


import logging as _logging
_log = _logging.getLogger("protactics.engines")


def _run_engines(db: Session, puerto_id: int, year: int, mes: int, actor_id=None):
    """Corre anomalías (mes) + SLA (puerto completo). Hace commit (vía SLA)."""
    anomalies.evaluar_mes(db, puerto_id, year, mes, actor_id=actor_id)
    sla_engine.evaluar_puerto(db, puerto_id, actor_id=actor_id)   # commit


def _run_engines_safe(db: Session, puerto_id: int, year: int, mes: int, actor_id=None):
    """Igual que _run_engines pero a prueba de fallos: nunca propaga la excepción
    (los datos cargados ya están confirmados antes de llamar a esto)."""
    try:
        _run_engines(db, puerto_id, year, mes, actor_id=actor_id)
    except Exception as e:
        db.rollback()
        _log.error("Motores de inteligencia operacional fallaron en "
                   "puerto=%s %s/%s: %s", puerto_id, year, mes, e)


def guardar_detalle(db: Session, puerto_id: int, year: int, mes: int,
                    raw_rows: list, filename: str, formato: str) -> int:
    """Persiste el DETALLE por escaneo (todas las columnas) + el índice de
    contenedores/matrículas, para los reportes que lo traen. Idempotente por día:
    reemplaza solo los días presentes en este archivo (igual que save_parsed_data),
    de modo que acumular reportes diarios no borra los días ya cargados. Devuelve
    el nº de filas de detalle guardadas. Para reportes-resumen (sin columnas de
    contenedor/matrícula) no hace nada y devuelve 0."""
    filas = identificadores.extraer_filas(raw_rows, year, mes)
    if not filas:
        return 0

    dias = {f["dia"] for f in filas if f["dia"] is not None}
    if dias:
        viejas = [r.id for r in db.query(EscaneoFila.id).filter(
            EscaneoFila.puerto_id == puerto_id, EscaneoFila.year == year,
            EscaneoFila.mes == mes, EscaneoFila.dia.in_(dias)).all()]
        if viejas:
            db.query(IndiceIdentificador).filter(
                IndiceIdentificador.fila_id.in_(viejas)).delete(synchronize_session=False)
            db.query(EscaneoFila).filter(
                EscaneoFila.id.in_(viejas)).delete(synchronize_session=False)

    for idx, f in enumerate(filas):
        fila = EscaneoFila(puerto_id=puerto_id, formato=formato, filename=filename,
                           fila_idx=idx, fecha_hora=f["fecha_hora"], year=year,
                           mes=mes, dia=f["dia"], datos=f["datos"],
                           cargado_en=datetime.utcnow())
        db.add(fila)
        db.flush()
        for c in f["contenedores"]:
            db.add(IndiceIdentificador(
                fila_id=fila.id, puerto_id=puerto_id, fecha_hora=f["fecha_hora"],
                tipo="contenedor", valor=c, valido=identificadores.validar_iso6346(c)))
        for valor, tipo_placa in f["placas"]:
            db.add(IndiceIdentificador(
                fila_id=fila.id, puerto_id=puerto_id, fecha_hora=f["fecha_hora"],
                tipo="placa", valor=valor, valido=None, tipo_placa=tipo_placa))
    db.commit()
    return len(filas)


def _guardar_detalle_safe(db: Session, puerto_id: int, year: int, mes: int,
                          raw_rows: list, filename: str, formato: str):
    """A prueba de fallos: la trazabilidad nunca debe romper la carga."""
    try:
        guardar_detalle(db, puerto_id, year, mes, raw_rows, filename, formato)
    except Exception as e:
        db.rollback()
        _log.error("Detalle de trazabilidad falló en puerto=%s %s/%s: %s",
                   puerto_id, year, mes, e)


def process_upload(db: Session, puerto: Puerto, year: int, mes: int,
                   raw_rows: list, filename: str) -> dict:
    """Detecta, parsea y guarda un archivo ya leído para (puerto, year, mes).

    Compartido por la carga individual y la carga masiva. NO valida permisos
    (eso es responsabilidad de quien llama). Lanza HTTPException 400 si el archivo
    no aporta escaneos para el período. Devuelve los totales del archivo y del mes.
    """
    fmt = detect_format(raw_rows)
    rows = rows_to_dicts(raw_rows) if fmt in ("standard", "tcbuen") else raw_rows

    month_name = MONTHS[mes - 1]
    data = parse_file(rows, puerto.nombre_corto, month_name, year, mes)

    if data["total_scans"] == 0:
        raise HTTPException(
            400,
            f"No se encontraron escaneos de {MONTHS[mes - 1]} {year} en este "
            f"archivo. Verifica que el archivo y el período sean correctos."
        )

    save_parsed_data(db, puerto.id, year, mes, data, filename)

    # Inteligencia operacional: detectar anomalías del mes y reevaluar el SLA del
    # puerto (racha + meses faltantes). Nunca debe romper la carga: si falla, se
    # revierte SOLO lo del motor (los datos ya están confirmados) y se registra.
    _run_engines_safe(db, puerto.id, year, mes)

    # Trazabilidad: guardar el detalle por escaneo (contenedores/matrículas) si el
    # reporte lo trae. A prueba de fallos (no rompe la carga).
    _guardar_detalle_safe(db, puerto.id, year, mes, raw_rows, filename, data["format"])

    month_total = db.query(func.coalesce(func.sum(EscaneosDiarios.total), 0))\
        .filter_by(puerto_id=puerto.id, year=year, mes=mes).scalar() or 0
    month_days = db.query(EscaneosDiarios)\
        .filter_by(puerto_id=puerto.id, year=year, mes=mes).count()

    return {
        "ok": True,
        "formato": data["format"],
        "total_escaneos": data["total_scans"],   # lo que aportó este archivo
        "dias_activos": data["days_active"],
        "pico_diario": data["peak_day"],
        "promedio_diario": data["avg_daily"],
        "total_mes": month_total,                # acumulado del mes
        "dias_mes": month_days,
    }


# ══════════════════════════════════════════════════════════════
#  ENDPOINTS
# ══════════════════════════════════════════════════════════════

# ── GET /puertos ────────────────────────────────────────────
@app.get("/puertos")
def get_puertos(year: Optional[int] = None,
                db: Session = Depends(get_db),
                user: User = Depends(get_current_user)):
    # Año en curso para los agregados de la lista del mapa (meses y escaneos
    # del año). Se puede sobrescribir con ?year=.
    if year is None:
        year = datetime.utcnow().year
    # Mes ANTERIOR al actual: la lista lateral muestra su disponibilidad.
    # En enero, el mes anterior es diciembre del año previo.
    now = datetime.utcnow()
    if now.month > 1:
        prev_y, prev_m = now.year, now.month - 1
    else:
        prev_y, prev_m = now.year - 1, 12
    q = db.query(Puerto).order_by(Puerto.id)
    ids = allowed_port_ids(user)        # None = todos
    if ids is not None:
        if not ids:                     # perfil con alcance sin puerto asignado
            return []
        q = q.filter(Puerto.id.in_(ids))
    puertos = q.all()
    result = []
    for p in puertos:
        # Sumar todos los escaneos del puerto
        total = sum(e.total for e in p.escaneos)
        meses = db.query(EscaneosDiarios.mes, EscaneosDiarios.year)\
                  .filter_by(puerto_id=p.id)\
                  .distinct().count()

        # ── Agregados del AÑO en curso (lista lateral del mapa) ──────────
        # Escaneos: suma de todos los escaneos diarios del año.
        escaneos_year = db.query(func.coalesce(func.sum(EscaneosDiarios.total), 0))\
            .filter_by(puerto_id=p.id, year=year).scalar() or 0
        # Meses cargados: número de meses del año con un archivo cargado.
        meses_year = db.query(ArchivosCargados)\
            .filter_by(puerto_id=p.id, year=year).count()
        # Disponibilidad del MES ANTERIOR (lo que muestra la lista lateral).
        disp_prev = db.query(Disponibilidad)\
            .filter_by(puerto_id=p.id, year=prev_y, mes=prev_m).first()
        disp_prev_val = disp_prev.valor if (disp_prev and disp_prev.valor is not None) else None

        # Disponibilidad más reciente
        last_avail = db.query(Disponibilidad)\
            .filter_by(puerto_id=p.id)\
            .order_by(Disponibilidad.year.desc(), Disponibilidad.mes.desc())\
            .first()
        result.append({
            "id": p.id, "nombre": p.nombre, "nombre_corto": p.nombre_corto,
            "departamento": p.departamento, "lat": p.lat, "lng": p.lng,
            "icono": p.icono, "sx": p.sx, "sy": p.sy, "formato": p.formato,
            "total_escaneos": total, "meses_cargados": meses,
            "year_actual": year,
            "escaneos_year": int(escaneos_year),
            "meses_year": meses_year,
            "mes_anterior": prev_m,
            "year_anterior": prev_y,
            "mes_anterior_nombre": MONTHS[prev_m - 1],
            "disponibilidad_mes_anterior": disp_prev_val,
            "ultima_disponibilidad": {
                "valor": last_avail.valor,
                "mes": last_avail.mes,
                "year": last_avail.year,
                "mes_nombre": MONTHS[last_avail.mes - 1]
            } if last_avail and last_avail.valor is not None else None
        })
    return result


# ── POST /upload/{puerto_id}/{year}/{mes} ───────────────────
@app.post("/upload/{puerto_id}/{year}/{mes}")
async def upload_file(
    puerto_id: int, year: int, mes: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not (1 <= mes <= 12):
        raise HTTPException(400, "Mes inválido")
    if not can_upload_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para cargar datos en este puerto")
    puerto = db.query(Puerto).filter_by(id=puerto_id).first()
    if not puerto:
        raise HTTPException(404, "Puerto no encontrado")

    content = await file.read()
    raw_rows = read_excel_rows(content, file.filename)

    # Validación de PUERTO: en una carga DIRIGIDA, el archivo debe pertenecer al
    # puerto elegido. Si su contenido/nombre apunta CLARAMENTE a otro puerto, se
    # rechaza en vez de archivar los datos en el puerto equivocado. Si la
    # detección es ambigua (None) se respeta la elección explícita del usuario.
    todos = db.query(Puerto).all()
    det_port = detect_port(raw_rows, file.filename, todos)
    if det_port is not None and det_port != puerto_id:
        otro = next((p for p in todos if p.id == det_port), None)
        nombre_otro = otro.nombre_corto if otro else f"#{det_port}"
        raise HTTPException(
            400,
            f"El archivo corresponde al puerto {nombre_otro}, no a "
            f"{puerto.nombre_corto}. Cárgalo en la tarjeta del puerto correcto."
        )

    # Validación de PERÍODO: en una carga DIRIGIDA a un mes concreto (botón o
    # arrastrar-y-soltar sobre la tarjeta), el archivo debe corresponder a ese
    # mes. Si su período dominante es otro, se rechaza con un error claro en vez
    # de archivar los datos en el mes equivocado.
    det_y, det_m, _ = detect_period(raw_rows)
    if det_y is not None and (det_y, det_m) != (year, mes):
        raise HTTPException(
            400,
            f"El archivo contiene datos de {MONTHS[det_m - 1]} {det_y}, no de "
            f"{MONTHS[mes - 1]} {year}. Suéltalo sobre la tarjeta del mes correcto."
        )

    result = process_upload(db, puerto, year, mes, raw_rows, file.filename)

    record_audit(accion="upload", entidad="escaneos",
                 entidad_id=f"{puerto_id}/{year}/{mes}", puerto_id=puerto_id,
                 actor=user, request=request,
                 detalle={"filename": file.filename, "formato": result["formato"],
                          "total_archivo": result["total_escaneos"],
                          "total_mes": result["total_mes"]})
    return result


# ── POST /upload/bulk ───────────────────────────────────────
@app.post("/upload/bulk")
async def upload_bulk(
    request: Request,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Carga masiva inteligente: hasta MAX_BULK_FILES archivos en una sola
    petición. Cada archivo se enruta a su (puerto, año, mes) leyéndolo, y se
    procesa de forma independiente: un archivo malo no aborta el lote. Cada
    archivo se confirma en su propia transacción (rollback aislado en error)."""
    if not files:
        raise HTTPException(400, "No se recibió ningún archivo")
    if len(files) > MAX_BULK_FILES:
        raise HTTPException(
            413, f"Máximo {MAX_BULK_FILES} archivos por lote (recibidos {len(files)})")

    puertos = db.query(Puerto).order_by(Puerto.id).all()
    by_id = {p.id: p for p in puertos}
    results = []
    summary = {"ok": 0, "error": 0, "needs_review": 0}

    for f in files:
        item = {"filename": f.filename, "status": "error", "puerto_id": None,
                "year": None, "mes": None, "total_escaneos": None, "message": None}
        try:
            content = await f.read()
            if len(content) > MAX_FILE_BYTES:
                item["message"] = f"Archivo demasiado grande (máx {MAX_FILE_BYTES // (1024*1024)} MB)"
                summary["error"] += 1
                results.append(item)
                continue

            raw_rows = read_excel_rows(content, f.filename)
            decision = route_file(raw_rows, f.filename, puertos)
            item.update(puerto_id=decision["puerto_id"], year=decision["year"],
                        mes=decision["mes"])

            # Enrutamiento dudoso → marcar para revisión, NO archivar a ciegas.
            if decision["confidence"] != "high":
                item["status"] = "needs_review"
                item["message"] = decision["reason"]
                item["routing"] = decision
                summary["needs_review"] += 1
                results.append(item)
                continue

            pid, yr, mo = decision["puerto_id"], decision["year"], decision["mes"]
            if not can_upload_port(user, pid):
                item["message"] = "No tienes permiso para cargar en este puerto"
                summary["error"] += 1
                results.append(item)
                continue

            puerto = by_id.get(pid)
            res = process_upload(db, puerto, yr, mo, raw_rows, f.filename)
            item.update(status="ok", total_escaneos=res["total_mes"],
                        message="Cargado", routing_source=decision["period_source"])
            if decision.get("multi_month"):
                item["multi_month"] = decision["multi_month"]
            summary["ok"] += 1
            results.append(item)

            record_audit(accion="upload_bulk", entidad="escaneos",
                         entidad_id=f"{pid}/{yr}/{mo}", puerto_id=pid,
                         actor=user, request=request,
                         detalle={"filename": f.filename, "formato": res["formato"],
                                  "total_mes": res["total_mes"],
                                  "period_source": decision["period_source"]})
        except HTTPException as e:
            db.rollback()        # aísla el fallo a este archivo
            item["message"] = e.detail
            summary["error"] += 1
            results.append(item)
        except Exception as e:
            db.rollback()
            item["message"] = f"Error al procesar: {e}"
            summary["error"] += 1
            results.append(item)

    return {"summary": summary, "total": len(files), "results": results}


# ── GET /data/{puerto_id}/{year}/{mes} ──────────────────────
@app.get("/data/{puerto_id}/{year}/{mes}")
def get_data(puerto_id: int, year: int, mes: int, db: Session = Depends(get_db),
             user: User = Depends(get_current_user)):
    if not can_view_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    puerto = db.query(Puerto).filter_by(id=puerto_id).first()
    if not puerto:
        raise HTTPException(404, "Puerto no encontrado")

    diarios = db.query(EscaneosDiarios)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes)\
        .order_by(EscaneosDiarios.dia).all()

    if not diarios:
        raise HTTPException(404, "Sin datos para ese período")

    horarios = db.query(EscaneosHorarios)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).all()

    ops = db.query(Operadores)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).all()

    avail = db.query(Disponibilidad)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).first()

    archivo = db.query(ArchivosCargados)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).first()

    # daily es por día; hora y operadores se acumulan SUMANDO entre los días.
    daily  = {str(d.dia): d.total for d in diarios}
    hourly = {}
    for h in horarios:
        hourly[str(h.hora)] = hourly.get(str(h.hora), 0) + h.total
    operators = {}
    for o in ops:
        operators[o.nombre] = operators.get(o.nombre, 0) + o.total

    total     = sum(d.total for d in diarios)
    days      = len(diarios)
    peak      = max(d.total for d in diarios)
    avg       = round(total / days) if days else 0

    return {
        "puerto_id": puerto_id,
        "puerto_nombre": puerto.nombre_corto,
        "year": year,
        "mes": mes,
        "mes_nombre": MONTHS[mes - 1],
        "total_scans": total,
        "days_active": days,
        "peak_day": peak,
        "avg_daily": avg,
        "daily": daily,
        "hourly": hourly,
        "operators": operators,
        "operator_count": len(operators),
        "disponibilidad": avail.valor if avail else None,
        "formato": archivo.formato if archivo else None,
    }


# ── GET /meses/{puerto_id} ──────────────────────────────────
@app.get("/meses/{puerto_id}")
def get_meses(puerto_id: int, db: Session = Depends(get_db),
              user: User = Depends(get_current_user)):
    """Retorna qué meses tienen datos para un puerto."""
    if not can_view_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    archivos = db.query(ArchivosCargados)\
        .filter_by(puerto_id=puerto_id)\
        .order_by(ArchivosCargados.year, ArchivosCargados.mes).all()

    disponibilidades = db.query(Disponibilidad)\
        .filter_by(puerto_id=puerto_id).all()
    disp_map = {(d.year, d.mes): d.valor for d in disponibilidades}

    out = []
    for a in archivos:
        est = sla_engine.clasificar_mes(db, puerto_id, a.year, a.mes)
        out.append({
            "year": a.year,
            "mes": a.mes,
            "mes_nombre": MONTHS[a.mes - 1],
            "total_escaneos": a.total_escaneos,
            "formato": a.formato,
            "cargado_en": a.cargado_en.isoformat(),
            "disponibilidad": disp_map.get((a.year, a.mes)),
            "estado_sla": {
                "estado": est["estado"], "motivo": est["motivo"],
                "observado": est["observado"], "umbral": est["umbral"],
                "dias_mantenimiento": est["dias_mantenimiento"],
            },
        })
    return out


# ── PUT /disponibilidad/{puerto_id}/{year}/{mes} ─────────────
class DispUpdate(BaseModel):
    valor: Optional[float] = None

@app.put("/disponibilidad/{puerto_id}/{year}/{mes}")
def set_disponibilidad(
    puerto_id: int, year: int, mes: int,
    body: DispUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    if not can_upload_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para editar este puerto")
    if body.valor is not None and not (0 <= body.valor <= 100):
        raise HTTPException(400, "Valor debe estar entre 0 y 100")

    existing = db.query(Disponibilidad)\
        .filter_by(puerto_id=puerto_id, year=year, mes=mes).first()
    valor_anterior = existing.valor if existing else None
    if existing:
        existing.valor = body.valor
        existing.actualizado = datetime.utcnow()
    else:
        db.add(Disponibilidad(puerto_id=puerto_id, year=year, mes=mes, valor=body.valor))
    db.commit()

    record_audit(accion="edit_disponibilidad", entidad="disponibilidad",
                 entidad_id=f"{puerto_id}/{year}/{mes}", puerto_id=puerto_id,
                 actor=user, request=request,
                 detalle={"antes": valor_anterior, "despues": body.valor})
    return {"ok": True, "puerto_id": puerto_id, "year": year, "mes": mes, "valor": body.valor}


# ── GET /disponibilidad/{puerto_id} ─────────────────────────
@app.get("/disponibilidad/{puerto_id}")
def get_disponibilidad(puerto_id: int, db: Session = Depends(get_db),
                       user: User = Depends(get_current_user)):
    if not can_view_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    items = db.query(Disponibilidad)\
        .filter_by(puerto_id=puerto_id)\
        .order_by(Disponibilidad.year.desc(), Disponibilidad.mes.desc()).all()
    return [{
        "year": d.year, "mes": d.mes,
        "mes_nombre": MONTHS[d.mes - 1],
        "valor": d.valor
    } for d in items]


# ── AUDITORÍA (solo admin) ──────────────────────────────────
@app.get("/api/audit/verify")
def audit_verify(db: Session = Depends(get_db), _admin: User = Depends(require_admin)):
    """Verifica la integridad de la cadena hash de la pista de auditoría.

    Devuelve {"ok": true, "count": n} si nadie alteró la pista, o
    {"ok": false, "broken_at": id, "count": n} si detecta manipulación.
    """
    return verify_chain(db)


@app.get("/api/audit")
def audit_list(limit: int = 100, offset: int = 0,
               db: Session = Depends(get_db), _admin: User = Depends(require_admin)):
    """Lista los eventos de auditoría más recientes (solo lectura, solo admin)."""
    limit = max(1, min(limit, 500))
    q = db.query(AuditLog).order_by(AuditLog.id.desc()).offset(max(0, offset)).limit(limit)
    return [{
        "id": a.id,
        "creado_en": a.creado_en.isoformat() if a.creado_en else None,
        "actor_email": a.actor_email,
        "actor_user_id": a.actor_user_id,
        "accion": a.accion,
        "entidad": a.entidad,
        "entidad_id": a.entidad_id,
        "puerto_id": a.puerto_id,
        "detalle": a.detalle,
        "ip": a.ip,
    } for a in q.all()]


# ══════════════════════════════════════════════════════════════
#  INTELIGENCIA OPERACIONAL — Mantenimiento, SLA y Alertas
# ══════════════════════════════════════════════════════════════

# ── Ventanas de mantenimiento ───────────────────────────────
MANT_TIPOS = ("programado", "falla_tecnica")


class MantCreate(BaseModel):
    tipo: str
    fecha_inicio: date
    fecha_fin: Optional[date] = None
    motivo: Optional[str] = None


class MantUpdate(BaseModel):
    tipo: Optional[str] = None
    fecha_inicio: Optional[date] = None
    fecha_fin: Optional[date] = None
    motivo: Optional[str] = None
    cerrar: Optional[bool] = None        # cerrar = fijar fecha_fin a hoy


def _mant_public(v: VentanaMantenimiento) -> dict:
    return {
        "id": v.id, "puerto_id": v.puerto_id, "tipo": v.tipo,
        "fecha_inicio": v.fecha_inicio.isoformat() if v.fecha_inicio else None,
        "fecha_fin": v.fecha_fin.isoformat() if v.fecha_fin else None,
        "motivo": v.motivo, "creada_por": v.creada_por,
        "creada_en": v.creada_en.isoformat() if v.creada_en else None,
        "cerrada_en": v.cerrada_en.isoformat() if v.cerrada_en else None,
        "abierta": v.fecha_fin is None,
    }


@app.get("/mantenimiento/{puerto_id}")
def get_mantenimiento(puerto_id: int, db: Session = Depends(get_db),
                      user: User = Depends(get_current_user)):
    if not can_view_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    items = db.query(VentanaMantenimiento).filter_by(puerto_id=puerto_id)\
        .order_by(VentanaMantenimiento.fecha_inicio.desc()).all()
    return [_mant_public(v) for v in items]


@app.post("/mantenimiento/{puerto_id}")
def create_mantenimiento(puerto_id: int, body: MantCreate, request: Request,
                         db: Session = Depends(get_db),
                         admin: User = Depends(require_admin)):
    if not db.query(Puerto).filter_by(id=puerto_id).first():
        raise HTTPException(404, "Puerto no encontrado")
    if body.tipo not in MANT_TIPOS:
        raise HTTPException(422, "Tipo inválido (programado | falla_tecnica)")
    if body.fecha_fin is not None and body.fecha_fin < body.fecha_inicio:
        raise HTTPException(422, "La fecha de fin no puede ser anterior a la de inicio")
    v = VentanaMantenimiento(
        puerto_id=puerto_id, tipo=body.tipo, fecha_inicio=body.fecha_inicio,
        fecha_fin=body.fecha_fin, motivo=(body.motivo or "").strip() or None,
        creada_por=admin.id, creada_en=datetime.utcnow(),
        cerrada_en=datetime.utcnow() if body.fecha_fin is not None else None,
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    record_audit(accion="create_mantenimiento", entidad="mantenimiento",
                 entidad_id=v.id, puerto_id=puerto_id, actor=admin, request=request,
                 detalle=_mant_public(v))
    return _mant_public(v)


@app.patch("/mantenimiento/{vid}")
def update_mantenimiento(vid: int, body: MantUpdate, request: Request,
                         db: Session = Depends(get_db),
                         admin: User = Depends(require_admin)):
    v = db.query(VentanaMantenimiento).filter_by(id=vid).first()
    if not v:
        raise HTTPException(404, "Ventana no encontrada")
    antes = _mant_public(v)
    if body.tipo is not None:
        if body.tipo not in MANT_TIPOS:
            raise HTTPException(422, "Tipo inválido (programado | falla_tecnica)")
        v.tipo = body.tipo
    if body.fecha_inicio is not None:
        v.fecha_inicio = body.fecha_inicio
    if body.cerrar:
        v.fecha_fin = date.today()
        v.cerrada_en = datetime.utcnow()
    elif body.fecha_fin is not None:
        v.fecha_fin = body.fecha_fin
        v.cerrada_en = datetime.utcnow()
    if body.motivo is not None:
        v.motivo = body.motivo.strip() or None
    if v.fecha_fin is not None and v.fecha_fin < v.fecha_inicio:
        raise HTTPException(422, "La fecha de fin no puede ser anterior a la de inicio")
    db.commit()
    db.refresh(v)
    record_audit(accion="update_mantenimiento", entidad="mantenimiento",
                 entidad_id=v.id, puerto_id=v.puerto_id, actor=admin, request=request,
                 detalle={"antes": antes, "despues": _mant_public(v)})
    return _mant_public(v)


@app.delete("/mantenimiento/{vid}")
def delete_mantenimiento(vid: int, request: Request, db: Session = Depends(get_db),
                         admin: User = Depends(require_admin)):
    v = db.query(VentanaMantenimiento).filter_by(id=vid).first()
    if not v:
        raise HTTPException(404, "Ventana no encontrada")
    snap = _mant_public(v)
    db.delete(v)
    db.commit()
    record_audit(accion="delete_mantenimiento", entidad="mantenimiento",
                 entidad_id=vid, puerto_id=snap["puerto_id"], actor=admin,
                 request=request, detalle=snap)
    return {"ok": True}


# ── Metas de SLA ────────────────────────────────────────────
class SLAUpdate(BaseModel):
    umbral: float
    metrica: str = "availability"
    periodo: str = "mensual"
    activo: bool = True


def _sla_public(s: SLA) -> dict:
    return {"id": s.id, "puerto_id": s.puerto_id, "metrica": s.metrica,
            "umbral": s.umbral, "periodo": s.periodo, "activo": s.activo}


@app.get("/sla")
def list_sla(db: Session = Depends(get_db), _admin: User = Depends(require_admin)):
    """Todas las metas configuradas (incluida la global, puerto_id NULL)."""
    return [_sla_public(s) for s in db.query(SLA).order_by(
        SLA.puerto_id.is_(None).desc(), SLA.puerto_id).all()]


@app.get("/sla/{puerto_id}/estado")
def sla_estado(puerto_id: int, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    """Meta efectiva + racha de incumplimiento del puerto (solo lectura)."""
    if not can_view_port(user, puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    meta = sla_engine.meta_efectiva(db, puerto_id)
    umbral = meta.umbral if meta else sla_engine.DEFAULT_AVAILABILITY_TARGET
    propia = meta is not None and meta.puerto_id == puerto_id
    return {
        "puerto_id": puerto_id,
        "umbral": umbral,
        "origen": "puerto" if propia else "global",
        "racha_incumplimiento": sla_engine.racha_incumplimiento(db, puerto_id),
        "racha_critica": sla_engine.RACHA_CRITICA,
    }


def _set_sla(db: Session, puerto_id, body: SLAUpdate):
    if body.metrica not in ("availability", "upload_deadline", "min_daily_scans"):
        raise HTTPException(422, "Métrica inválida")
    if body.metrica == "availability" and not (0 <= body.umbral <= 100):
        raise HTTPException(422, "El umbral de disponibilidad debe estar entre 0 y 100")
    row = db.query(SLA).filter_by(puerto_id=puerto_id, metrica=body.metrica).first()
    antes = _sla_public(row) if row else None
    if row:
        row.umbral = body.umbral
        row.periodo = body.periodo
        row.activo = body.activo
    else:
        row = SLA(puerto_id=puerto_id, metrica=body.metrica, umbral=body.umbral,
                  periodo=body.periodo, activo=body.activo)
        db.add(row)
    db.commit()
    db.refresh(row)
    return row, antes


@app.put("/sla/global")
def set_sla_global(body: SLAUpdate, request: Request, db: Session = Depends(get_db),
                   admin: User = Depends(require_admin)):
    """Fija la meta por defecto global (puerto_id NULL)."""
    row, antes = _set_sla(db, None, body)
    record_audit(accion="edit_sla", entidad="sla", entidad_id=row.id,
                 puerto_id=None, actor=admin, request=request,
                 detalle={"antes": antes, "despues": _sla_public(row)})
    return _sla_public(row)


@app.put("/sla/{puerto_id}")
def set_sla_puerto(puerto_id: int, body: SLAUpdate, request: Request,
                   db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Fija/actualiza la meta de un puerto concreto."""
    if not db.query(Puerto).filter_by(id=puerto_id).first():
        raise HTTPException(404, "Puerto no encontrado")
    row, antes = _set_sla(db, puerto_id, body)
    record_audit(accion="edit_sla", entidad="sla", entidad_id=row.id,
                 puerto_id=puerto_id, actor=admin, request=request,
                 detalle={"antes": antes, "despues": _sla_public(row)})
    return _sla_public(row)


# ── Alertas ─────────────────────────────────────────────────
def _alerta_public(a: Alerta) -> dict:
    return {
        "id": a.id, "puerto_id": a.puerto_id, "tipo": a.tipo,
        "severidad": a.severidad, "mensaje": a.mensaje, "estado": a.estado,
        "year": a.year, "mes": a.mes, "dia": a.dia,
        "mes_nombre": MONTHS[a.mes - 1] if a.mes else None,
        "payload": a.payload,
        "creada_en": a.creada_en.isoformat() if a.creada_en else None,
        "resuelta_en": a.resuelta_en.isoformat() if a.resuelta_en else None,
    }


_SEV_ORDEN = {"critical": 0, "warning": 1, "info": 2}


@app.get("/alertas")
def list_alertas(puerto_id: Optional[int] = None, estado: Optional[str] = None,
                 tipo: Optional[str] = None, severidad: Optional[str] = None,
                 limit: int = 200, db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    q = db.query(Alerta)
    ids = allowed_port_ids(user)
    if ids is not None:
        if not ids:
            return []
        q = q.filter(Alerta.puerto_id.in_(ids))
    if puerto_id is not None:
        if not can_view_port(user, puerto_id):
            raise HTTPException(403, "No tienes permiso para ver este puerto")
        q = q.filter(Alerta.puerto_id == puerto_id)
    if estado:
        q = q.filter(Alerta.estado == estado)
    if tipo:
        q = q.filter(Alerta.tipo == tipo)
    if severidad:
        q = q.filter(Alerta.severidad == severidad)
    rows = q.all()
    rows.sort(key=lambda a: (_SEV_ORDEN.get(a.severidad, 9),
                             -(a.creada_en.timestamp() if a.creada_en else 0)))
    return [_alerta_public(a) for a in rows[:max(1, min(limit, 500))]]


@app.get("/alertas/resumen")
def resumen_alertas(db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """Conteo de alertas ABIERTAS por puerto y severidad (insignias del mapa)."""
    q = db.query(Alerta).filter(Alerta.estado.in_(("open", "acknowledged")))
    ids = allowed_port_ids(user)
    if ids is not None:
        if not ids:
            return {}
        q = q.filter(Alerta.puerto_id.in_(ids))
    out = {}
    for a in q.all():
        d = out.setdefault(a.puerto_id, {"critical": 0, "warning": 0, "info": 0})
        d[a.severidad] = d.get(a.severidad, 0) + 1
    return out


def _get_alerta_mutable(aid: int, db: Session, user: User) -> Alerta:
    a = db.query(Alerta).filter_by(id=aid).first()
    if not a:
        raise HTTPException(404, "Alerta no encontrada")
    if not can_manage_alerts(user, a.puerto_id):
        raise HTTPException(403, "No tienes permiso para gestionar alertas de este puerto")
    return a


@app.post("/alertas/{aid}/ack")
def ack_alerta(aid: int, request: Request, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    a = _get_alerta_mutable(aid, db, user)
    if a.estado == "open":
        a.estado = "acknowledged"
        db.commit()
        record_audit(accion="ack_alerta", entidad="alertas", entidad_id=a.id,
                     puerto_id=a.puerto_id, actor=user, request=request,
                     detalle={"tipo": a.tipo})
    return _alerta_public(a)


@app.post("/alertas/{aid}/resolve")
def resolve_alerta(aid: int, request: Request, db: Session = Depends(get_db),
                   user: User = Depends(get_current_user)):
    a = _get_alerta_mutable(aid, db, user)
    if a.estado != "resolved":
        a.estado = "resolved"
        a.resuelta_en = datetime.utcnow()
        a.resuelta_por = user.id
        db.commit()
        record_audit(accion="resolve_alerta", entidad="alertas", entidad_id=a.id,
                     puerto_id=a.puerto_id, actor=user, request=request,
                     detalle={"tipo": a.tipo})
    return _alerta_public(a)


@app.post("/alertas/recalcular/{puerto_id}/{year}/{mes}")
def recalcular(puerto_id: int, year: int, mes: int, request: Request,
               db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Reejecuta anomalías (mes) + SLA (puerto) tras cambiar metas o mantenimiento."""
    if not (1 <= mes <= 12):
        raise HTTPException(400, "Mes inválido")
    if not db.query(Puerto).filter_by(id=puerto_id).first():
        raise HTTPException(404, "Puerto no encontrado")
    _run_engines(db, puerto_id, year, mes, actor_id=admin.id)
    record_audit(accion="recalcular_alertas", entidad="alertas",
                 entidad_id=f"{puerto_id}/{year}/{mes}", puerto_id=puerto_id,
                 actor=admin, request=request)
    abiertas = db.query(Alerta).filter(
        Alerta.puerto_id == puerto_id, Alerta.estado.in_(("open", "acknowledged"))
    ).count()
    return {"ok": True, "puerto_id": puerto_id, "year": year, "mes": mes,
            "alertas_abiertas": abiertas}


# ══════════════════════════════════════════════════════════════
#  BÚSQUEDA Y TRAZABILIDAD DE CONTENEDORES / VEHÍCULOS
# ══════════════════════════════════════════════════════════════
def _nombres_puertos(db: Session) -> dict:
    return {p.id: p.nombre_corto for p in db.query(Puerto).all()}


def _puertos_con_detalle(db: Session) -> set:
    """IDs de puertos que han enviado reporte de DETALLE (rastreables)."""
    return {r[0] for r in db.query(EscaneoFila.puerto_id).distinct().all()}


def _ident_publico(ix, nombres) -> dict:
    return {
        "fila_id": ix.fila_id, "puerto_id": ix.puerto_id,
        "puerto": nombres.get(ix.puerto_id),
        "tipo": ix.tipo, "valor": ix.valor, "valido": ix.valido,
        "tipo_placa": ix.tipo_placa,
        "fecha_hora": ix.fecha_hora.isoformat() if ix.fecha_hora else None,
    }


def _aplicar_alcance(query, user):
    """Filtra una query de IndiceIdentificador por los puertos visibles del user.
    Devuelve (query, vacio) — vacio=True si el user no puede ver ningún puerto."""
    ids = allowed_port_ids(user)
    if ids is not None:
        if not ids:
            return query, True
        query = query.filter(IndiceIdentificador.puerto_id.in_(ids))
    return query, False


@app.get("/buscar")
def buscar(q: str, request: Request, tipo: str = "auto",
           puerto_id: Optional[int] = None, solo_validos: bool = False,
           limit: int = 200, db: Session = Depends(get_db),
           user: User = Depends(get_current_user)):
    """Busca un contenedor o matrícula (match exacto, normalizado). Respeta el
    alcance de puertos del usuario y audita la consulta."""
    valor = identificadores.normalizar(q)
    if len(valor) < 3:
        raise HTTPException(400, "La búsqueda requiere al menos 3 caracteres.")
    if tipo not in ("auto", "contenedor", "placa"):
        raise HTTPException(422, "tipo inválido (auto|contenedor|placa)")

    # Filtros comunes (todo menos el valor). La condición sobre `valor` se aplica
    # después en cascada: exacto → prefijo → contiene, para que escribir un
    # fragmento ("FFAU") encuentre el contenedor completo ("FFAU5573878").
    base = db.query(IndiceIdentificador)
    if tipo in ("contenedor", "placa"):
        base = base.filter(IndiceIdentificador.tipo == tipo)
    if solo_validos:
        base = base.filter(IndiceIdentificador.valido.is_(True))
    if puerto_id is not None:
        if not can_view_port(user, puerto_id):
            raise HTTPException(403, "No tienes permiso para ver este puerto")
        base = base.filter(IndiceIdentificador.puerto_id == puerto_id)
    base, vacio = _aplicar_alcance(base, user)

    rows, modo = [], "exacto"
    if not vacio:
        rows = base.filter(IndiceIdentificador.valor == valor).all()
        if not rows:
            rows = base.filter(IndiceIdentificador.valor.like(valor + "%")).all()
            modo = "prefijo"
        if not rows:
            rows = base.filter(IndiceIdentificador.valor.like("%" + valor + "%")).all()
            modo = "contiene"

    rows.sort(key=lambda r: (r.valor, r.fecha_hora or datetime.min))
    nombres = _nombres_puertos(db)
    resultados = [_ident_publico(r, nombres) for r in rows[:max(1, min(limit, 1000))]]

    record_audit(accion="buscar_identificador", entidad="busqueda",
                 entidad_id=valor, actor=user, request=request,
                 detalle={"tipo": tipo, "modo": modo, "resultados": len(rows)})
    return {"q": valor, "tipo": tipo, "modo": modo,
            "total": len(rows), "resultados": resultados}


@app.get("/contenedor/{num}/trayecto")
def trayecto_contenedor(num: str, request: Request,
                        db: Session = Depends(get_db),
                        user: User = Depends(get_current_user)):
    """Itinerario de un contenedor: puertos por los que pasó, en orden cronológico.
    Incluye advertencia de cobertura (puertos sin reporte de detalle no aparecen
    aunque el contenedor pudiera haber pasado por ellos)."""
    valor = identificadores.normalizar(num)
    if len(valor) < 3:
        raise HTTPException(400, "El contenedor requiere al menos 3 caracteres.")

    query = db.query(IndiceIdentificador).filter(
        IndiceIdentificador.tipo == "contenedor", IndiceIdentificador.valor == valor)
    query, vacio = _aplicar_alcance(query, user)
    rows = [] if vacio else query.all()
    rows.sort(key=lambda r: (r.fecha_hora or datetime.min))

    nombres = _nombres_puertos(db)
    pasos = [_ident_publico(r, nombres) for r in rows]

    # Puertos distintos en orden de primera aparición.
    puertos, vistos = [], set()
    for r in rows:
        if r.puerto_id not in vistos:
            vistos.add(r.puerto_id)
            puertos.append({"puerto_id": r.puerto_id,
                            "puerto": nombres.get(r.puerto_id),
                            "primera_fecha": r.fecha_hora.isoformat() if r.fecha_hora else None})

    con_detalle = _puertos_con_detalle(db)
    ids_visibles = allowed_port_ids(user)
    todos = [p.id for p in db.query(Puerto).all()
             if ids_visibles is None or p.id in ids_visibles]
    sin_cobertura = [{"puerto_id": pid, "puerto": nombres.get(pid)}
                     for pid in todos if pid not in con_detalle]

    record_audit(accion="trayecto_contenedor", entidad="busqueda",
                 entidad_id=valor, actor=user, request=request,
                 detalle={"puertos": len(puertos), "pasos": len(rows)})
    return {"contenedor": valor, "valido": identificadores.validar_iso6346(valor),
            "puertos": puertos, "pasos": pasos,
            "cobertura": {"puertos_sin_detalle": sin_cobertura}}


@app.get("/escaneo/{fila_id}")
def escaneo_detalle(fila_id: int, db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """Ficha completa de un escaneo: TODAS las columnas guardadas + sus
    identificadores (contenedores/matrículas)."""
    fila = db.query(EscaneoFila).filter_by(id=fila_id).first()
    if not fila:
        raise HTTPException(404, "Escaneo no encontrado")
    if not can_view_port(user, fila.puerto_id):
        raise HTTPException(403, "No tienes permiso para ver este puerto")
    nombres = _nombres_puertos(db)
    idents = db.query(IndiceIdentificador).filter_by(fila_id=fila.id).all()
    return {
        "id": fila.id, "puerto_id": fila.puerto_id,
        "puerto": nombres.get(fila.puerto_id), "formato": fila.formato,
        "filename": fila.filename, "year": fila.year, "mes": fila.mes, "dia": fila.dia,
        "fecha_hora": fila.fecha_hora.isoformat() if fila.fecha_hora else None,
        "datos": fila.datos or {},
        "contenedores": [i.valor for i in idents if i.tipo == "contenedor"],
        "placas": [{"valor": i.valor, "tipo": i.tipo_placa}
                   for i in idents if i.tipo == "placa"],
    }
